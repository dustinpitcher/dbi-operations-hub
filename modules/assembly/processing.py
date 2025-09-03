import pandas as pd
import numpy as np
from typing import Dict, List, Any
import os
from datetime import datetime


class AssemblyProcessor:
    """
    Core business logic for DBI Assembly Generation System
    Processes inventory, sales, and BOM data to generate assembly and transfer recommendations
    """
    
    def __init__(self, debug_mode=False):
        self.debug_mode = debug_mode  # Control debug logging for performance  
        self.availability_df = None
        self.kpi_df = None
        self.bom_df = None
        self.TARGET_DAYS_INVENTORY = 30  # Goal: 30 days of inventory
    
    def load_data(self, availability_path: str, replenishment_path: str, bom_path: str):
        """Load and clean all data files"""
        try:
            print("Loading data files...")
            
            # Load availability report
            self.availability_df = pd.read_csv(availability_path)
            print(f"Loaded availability data: {len(self.availability_df)} records")
            
            # Load replenishment/sales data (better than KPI data)
            self.kpi_df = pd.read_csv(replenishment_path)
            print(f"Loaded replenishment data: {len(self.kpi_df)} records")
            
            # Load BOM data from Excel (headers are in row 3)
            self.bom_df = pd.read_excel(bom_path, header=2)  # 0-indexed, so row 3 = header=2
            print(f"Loaded BOM data: {len(self.bom_df)} records")
            
            # Clean and prepare data
            self._clean_data()
            
        except Exception as e:
            print(f"Error loading data: {e}")
            raise
    
    def _clean_data(self):
        """Clean and standardize data"""
        
        # Clean availability data
        self.availability_df['OnHand'] = pd.to_numeric(self.availability_df['OnHand'], errors='coerce').fillna(0)
        self.availability_df['Available'] = pd.to_numeric(self.availability_df['Available'], errors='coerce').fillna(0)
        
        # Clean replenishment data - handle the quoted SKU format  
        if 'SKU' in self.kpi_df.columns:
            # Remove quotes and equals sign from SKU format (="1590" or "1348 becomes 1590, 1348)
            self.kpi_df['SKU'] = (self.kpi_df['SKU'].astype(str)
                                 .str.replace('=', '', regex=False)  # Remove equals
                                 .str.strip('"')  # Remove quotes from start/end
                                 .str.strip("'")  # Remove single quotes 
                                 .str.replace('"', '', regex=False))  # Remove any remaining quotes
        
        # Clean sales data
        if 'AVG sales/mo' in self.kpi_df.columns:
            self.kpi_df['AVG sales/mo'] = pd.to_numeric(self.kpi_df['AVG sales/mo'], errors='coerce').fillna(0)
        
        # Clean BOM data
        self.bom_df['Quantity'] = pd.to_numeric(self.bom_df['Quantity'], errors='coerce').fillna(0)
        self.bom_df['Available'] = pd.to_numeric(self.bom_df['Available'], errors='coerce').fillna(0)
        
        print("Data cleaning completed")
    
    def analyze_assembly_capacity(self) -> Dict[str, Any]:
        """
        Analyze which products can be assembled and which cannot
        Returns: Dictionary with assembly_ready, cannot_assemble, transfer_recommendations, and summary
        """
        assembly_ready = []
        cannot_assemble = []
        
        print("Analyzing assembly capacity...")
        
        # Step 1: Start with all products that have BOMs (can be assembled)
        bom_products = self.bom_df['Product SKU'].unique()
        print(f"Found {len(bom_products)} unique products with BOMs")
        
        # Pre-filter availability and sales data to only relevant SKUs (major speedup)
        availability_filtered = self.availability_df[self.availability_df['SKU'].isin(bom_products)]
        kpi_filtered = self.kpi_df[self.kpi_df['SKU'].isin(bom_products)]
        
        if self.debug_mode:
            print(f"Pre-filtered availability: {len(availability_filtered)} from {len(self.availability_df)} records")
            print(f"Pre-filtered sales data: {len(kpi_filtered)} from {len(self.kpi_df)} records")
        
        # Step 2: For each BOM product, check inventory status and component availability
        for product_sku in bom_products:
            # Get current inventory in NC-Main locations (using pre-filtered data)
            current_stock = availability_filtered[
                (availability_filtered['SKU'] == product_sku) &
                (availability_filtered['Location'].str.contains('NC', na=False)) &
                (availability_filtered['Location'].str.contains('Main', na=False))
            ]['Available'].sum()
            
            # Get product name from availability data
            product_name_data = self.availability_df[self.availability_df['SKU'] == product_sku]
            product_name = 'Unknown'
            if not product_name_data.empty:
                product_name = product_name_data['ProductName'].iloc[0] if 'ProductName' in product_name_data.columns else 'Unknown'
            
            # Get sales velocity data (using pre-filtered data)
            sales_data = kpi_filtered[kpi_filtered['SKU'] == product_sku]
            avg_monthly_sales = 0
            daily_sales = 0
            days_of_inventory = float('inf')
            
            if not sales_data.empty:
                avg_monthly_sales = pd.to_numeric(sales_data['AVG sales/mo'].iloc[0], errors='coerce')
                if not pd.isna(avg_monthly_sales) and avg_monthly_sales > 0:
                    daily_sales = avg_monthly_sales / 30
                    days_of_inventory = current_stock / daily_sales if daily_sales > 0 else float('inf')
            
            # Step 3: Check if this product needs assembly (low stock OR no sales data but zero inventory)
            needs_assembly = False
            if avg_monthly_sales > 0:
                # Product with sales data - check if below target days
                needs_assembly = days_of_inventory < self.TARGET_DAYS_INVENTORY
            elif current_stock <= 5:
                # Product without sales data but very low/zero stock - might still be valuable to assemble
                needs_assembly = True
                avg_monthly_sales = 1  # Assume minimal sales for calculation
                daily_sales = avg_monthly_sales / 30
            
            if not needs_assembly:
                continue
            
            if self.debug_mode:
                print(f"Analyzing {product_sku}: {current_stock} stock, {days_of_inventory:.1f} days inventory")
            
            # Step 4: Get BOM and aggregate component requirements
            product_bom = self.bom_df[self.bom_df['Product SKU'] == product_sku]
            component_requirements = {}
            
            for _, bom_entry in product_bom.iterrows():
                component_sku = bom_entry['Component SKU']
                required_qty = pd.to_numeric(bom_entry['Quantity'], errors='coerce')
                
                if pd.isna(required_qty) or required_qty <= 0:
                    continue
                
                # Skip SV components (service/non-inventory items)
                if str(component_sku).startswith('SV'):
                    continue
                
                # Aggregate requirements (sum up multiple BOM entries for same component)
                if component_sku in component_requirements:
                    component_requirements[component_sku] += required_qty
                else:
                    component_requirements[component_sku] = required_qty
            
            if not component_requirements:
                continue
            
            # Step 5: Check component availability
            can_assemble = True
            min_assemblies = float('inf')
            missing_components = []
            
            for component_sku, required_qty in component_requirements.items():
                # Get total available quantity for this component across all locations
                component_availability = self.availability_df[
                    self.availability_df['SKU'] == component_sku
                ]['Available'].sum()
                
                if component_availability < required_qty:
                    can_assemble = False
                    missing_components.append({
                        'sku': component_sku,
                        'required': required_qty,
                        'available': component_availability,
                        'shortage': required_qty - component_availability
                    })
                else:
                    # Calculate how many complete assemblies we can make
                    possible_assemblies = int(component_availability // required_qty)
                    min_assemblies = min(min_assemblies, possible_assemblies)
            
            # Step 6: Generate recommendations
            if can_assemble and min_assemblies > 0 and min_assemblies != float('inf'):
                # Calculate recommended assembly quantity
                if daily_sales > 0:
                    needed_inventory = (self.TARGET_DAYS_INVENTORY * daily_sales) - current_stock
                    recommended_assembly = min(max(1, int(needed_inventory)), min_assemblies)
                else:
                    # For products without sales data, recommend a conservative amount
                    recommended_assembly = min(10, min_assemblies)
                
                if recommended_assembly > 0:
                    assembly_ready.append({
                        'product_sku': product_sku,
                        'product_name': product_name,
                        'quantity_for_assembly': recommended_assembly,
                        'available_in_nc': int(current_stock),
                        'avg_monthly_sales': round(avg_monthly_sales, 2),
                        'days_of_inventory': round(days_of_inventory, 1) if days_of_inventory != float('inf') else 0,
                        'max_possible_assemblies': min_assemblies,
                        'components_needed': len(component_requirements)
                    })
            else:
                # Product cannot be assembled due to missing components
                cannot_assemble.append({
                    'product_sku': product_sku,
                    'product_name': product_name,
                    'missing_components': [comp['sku'] for comp in missing_components],
                    'component_details': missing_components,
                    'total_components_required': len(component_requirements),
                    'missing_components_count': len(missing_components)
                })
        
        print(f"Assembly analysis complete:")
        print(f"  - {len(assembly_ready)} products ready for assembly")
        print(f"  - {len(cannot_assemble)} products cannot be assembled")
        
        # Enhanced sorting: prioritize by urgency, then quantity, then potential
        def assembly_priority(item):
            # Priority 1: Urgency (negative stock = highest priority)
            urgency_score = 1000 if item['available_in_nc'] < 0 else (
                500 if item['available_in_nc'] == 0 else (
                    100 if item['days_of_inventory'] < 5 else (
                        50 if item['days_of_inventory'] < 15 else 10
                    )
                )
            )
            
            # Priority 2: Assembly quantity needed
            quantity_score = item['quantity_for_assembly']
            
            # Priority 3: Max possible assemblies (tie-breaker)
            potential_score = item['max_possible_assemblies'] * 0.1
            
            return urgency_score + quantity_score + potential_score
        
        assembly_ready.sort(key=assembly_priority, reverse=True)
        
        # Add avg_monthly_sales to cannot_assemble items
        for item in cannot_assemble:
            product_sku = item['product_sku']
            # Get sales data for this SKU
            sales_data = self.kpi_df[self.kpi_df['SKU'] == product_sku]
            if not sales_data.empty:
                avg_sales = sales_data['AVG sales/mo'].iloc[0]
                # Convert from string if needed and handle NaN
                try:
                    item['avg_monthly_sales'] = float(avg_sales) if pd.notna(avg_sales) else 0
                except (ValueError, TypeError):
                    item['avg_monthly_sales'] = 0
            else:
                item['avg_monthly_sales'] = 0
        
        # Return as dictionary for easier access
        return {
            'assembly_ready': assembly_ready,
            'cannot_assemble': cannot_assemble,
            'transfer_recommendations': self.analyze_transfer_needs(),
            'summary': {
                'total_products_analyzed': len(bom_products),
                'assembly_ready': len(assembly_ready),
                'cannot_assemble': len(cannot_assemble),
                'transfer_recommendations': len(self.analyze_transfer_needs())
            }
        }
    
    def analyze_transfer_needs(self) -> List[Dict]:
        """
        Identify items in NC-Armory that should be transferred to NC-Main
        Based on sales velocity and current NC-Main inventory
        """
        transfer_recommendations = []
        
        # Get items in NC-Armory
        armory_items = self.availability_df[
            self.availability_df['Location'] == 'NC - Armory'
        ].copy()
        
        for _, item in armory_items.iterrows():
            sku = item['SKU']
            armory_qty = item['Available']
            product_name = item.get('ProductName', 'Unknown')
            
            if armory_qty <= 0:
                continue
            
            # Get quantity in NC-Main
            main_qty = self.availability_df[
                (self.availability_df['SKU'] == sku) & 
                (self.availability_df['Location'] == 'NC - Main')
            ]['Available'].sum()
            
            # Get sales velocity
            sales_data = self.kpi_df[self.kpi_df['SKU'] == sku]
            avg_monthly_sales = sales_data['AVG sales/mo'].iloc[0] if not sales_data.empty else 0
            
            if avg_monthly_sales > 0:
                daily_sales = avg_monthly_sales / 30
                days_supply_in_main = main_qty / daily_sales if daily_sales > 0 else float('inf')
                
                # If NC-Main has less than 30 days supply, recommend transfer
                if days_supply_in_main < self.TARGET_DAYS_INVENTORY:
                    needed_qty = (self.TARGET_DAYS_INVENTORY * daily_sales) - main_qty
                    suggested_transfer = min(int(needed_qty), int(armory_qty))
                    
                    if suggested_transfer > 0:
                        transfer_recommendations.append({
                            'sku': sku,
                            'product_name': product_name,
                            'qty_in_armory': int(armory_qty),
                            'qty_in_main': int(main_qty),
                            'avg_monthly_sales': round(avg_monthly_sales, 2),
                            'days_supply_in_main': round(days_supply_in_main, 1),
                            'suggested_transfer': suggested_transfer
                        })
            else:
                # For items with no sales data, consider transfer if Main has 0
                if main_qty == 0 and armory_qty > 0:
                    transfer_recommendations.append({
                        'sku': sku,
                        'product_name': product_name,
                        'qty_in_armory': int(armory_qty),
                        'qty_in_main': int(main_qty),
                        'avg_monthly_sales': 0,
                        'days_supply_in_main': 0,
                        'suggested_transfer': min(10, int(armory_qty))  # Conservative transfer
                    })
        
        # Sort by priority (lowest days supply first, then highest sales)
        transfer_recommendations.sort(key=lambda x: (x['days_supply_in_main'], -x['avg_monthly_sales']))
        
        return transfer_recommendations
    
    def export_to_excel(self, results: Dict[str, Any], output_path: str = "DBI_Assembly_Reports.xlsx"):
        """
        Export results to Excel file with separate sheets for each report
        Enhanced with sorting by average monthly sales and professional styling
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # Sheet 1: Assembly Ready Products
            assembly_df = pd.DataFrame(results['assembly_ready'])
            if not assembly_df.empty:
                # Add product names from KPI data
                assembly_df['product_name'] = assembly_df['product_sku'].apply(
                    lambda sku: self._get_product_name(sku)
                )
                
                # Sort by average monthly sales (highest first)
                assembly_df = assembly_df.sort_values('avg_monthly_sales', ascending=False)
                
                # Rename columns for better presentation
                column_mapping = {
                    'product_sku': 'Product SKU',
                    'product_name': 'Product Name',
                    'quantity_for_assembly': 'Qty for Assembly', 
                    'available_in_nc': 'Available in NC',
                    'avg_monthly_sales': 'Avg Monthly Sales',
                    'days_of_inventory': 'Days of Inventory',
                    'max_possible_assemblies': 'Max Possible Assemblies',
                    'components_needed': 'Components Needed'
                }
                assembly_df = assembly_df.rename(columns=column_mapping)
                assembly_df.to_excel(writer, sheet_name='Assembly Ready', index=False)
                
                # Apply professional styling
                worksheet = writer.sheets['Assembly Ready']
                self._style_worksheet(worksheet, assembly_df)
            
            # Sheet 2: Cannot Assemble
            cannot_df = pd.DataFrame(results['cannot_assemble'])
            if not cannot_df.empty:
                # Ensure avg_monthly_sales exists for sorting
                for i, row in cannot_df.iterrows():
                    if 'avg_monthly_sales' not in row or pd.isna(row.get('avg_monthly_sales', 0)):
                        product_sku = row['product_sku']
                        sales_data = self.kpi_df[self.kpi_df['SKU'] == product_sku]
                        avg_sales = 0
                        if not sales_data.empty:
                            avg_sales = pd.to_numeric(sales_data['AVG sales/mo'].iloc[0], errors='coerce')
                            if pd.isna(avg_sales):
                                avg_sales = 0
                        cannot_df.at[i, 'avg_monthly_sales'] = avg_sales
                
                # Sort by average monthly sales (highest first)
                cannot_df = cannot_df.sort_values('avg_monthly_sales', ascending=False)
                
                # Expand missing components for better readability
                expanded_cannot = []
                for _, row in cannot_df.iterrows():
                    missing_components = row['missing_components'] if isinstance(row['missing_components'], list) else [row['missing_components']]
                    
                    # Get component shortage details if available
                    shortage_details = []
                    if 'component_details' in row and row['component_details']:
                        for comp in row['component_details'][:5]:  # Limit to first 5
                            if isinstance(comp, dict):
                                shortage_details.append(f"{comp['sku']} (need {comp['shortage']} more)")
                            else:
                                shortage_details.append(str(comp))
                    
                    expanded_cannot.append({
                        'Product SKU': row['product_sku'],
                        'Product Name': row.get('product_name', 'Unknown'),
                        'Avg Monthly Sales': round(row.get('avg_monthly_sales', 0), 2),
                        'Missing Components': ', '.join(missing_components[:5]),  # Limit to first 5
                        'Shortage Details': ', '.join(shortage_details) if shortage_details else 'N/A',
                        'Total Components Required': row.get('total_components_required', 0),
                        'Missing Components Count': row.get('missing_components_count', len(missing_components))
                    })
                
                cannot_formatted_df = pd.DataFrame(expanded_cannot)
                cannot_formatted_df.to_excel(writer, sheet_name='Cannot Assemble', index=False)
                
                # Apply professional styling
                worksheet = writer.sheets['Cannot Assemble']
                self._style_worksheet(worksheet, cannot_formatted_df)
            
            # Sheet 3: Transfer Recommendations
            transfer_df = pd.DataFrame(results['transfer_recommendations'])
            if not transfer_df.empty:
                # Sort by average monthly sales (highest first)
                transfer_df = transfer_df.sort_values('avg_monthly_sales', ascending=False)
                
                # Rename columns for better presentation
                transfer_column_mapping = {
                    'sku': 'SKU',
                    'product_name': 'Product Name',
                    'qty_in_armory': 'Qty in Armory',
                    'qty_in_main': 'Qty in Main',
                    'avg_monthly_sales': 'Avg Monthly Sales',
                    'days_supply_in_main': 'Days Supply in Main',
                    'suggested_transfer': 'Suggested Transfer'
                }
                transfer_df = transfer_df.rename(columns=transfer_column_mapping)
                transfer_df.to_excel(writer, sheet_name='Transfer Recommendations', index=False)
                
                # Apply professional styling
                worksheet = writer.sheets['Transfer Recommendations']
                self._style_worksheet(worksheet, transfer_df)
            
            # Sheet 4: Summary Dashboard
            summary_data = [
                ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Total Products Analyzed', results['summary']['total_products_analyzed']],
                ['Products Ready for Assembly', results['summary']['assembly_ready']],
                ['Products Cannot Assemble', results['summary']['cannot_assemble']],
                ['Transfer Recommendations', results['summary']['transfer_recommendations']],
                ['Target Days Inventory', self.TARGET_DAYS_INVENTORY]
            ]
            
            summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary sheet
            worksheet = writer.sheets['Summary']
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 20
            
            # Apply styling to summary sheet
            worksheet = writer.sheets['Summary']
            self._style_worksheet(worksheet, summary_df, is_summary=True)
            
            # Ensure at least one sheet is visible and active (fix Excel error)
            if 'Assembly Ready' in writer.sheets:
                writer.book.active = writer.book['Assembly Ready']
            elif 'Cannot Assemble' in writer.sheets:
                writer.book.active = writer.book['Cannot Assemble']
            elif 'Transfer Recommendations' in writer.sheets:
                writer.book.active = writer.book['Transfer Recommendations']
            else:
                writer.book.active = writer.book['Summary']
        
        print(f"Excel report exported to: {output_path}")
        return output_path
    
    def _get_product_name(self, product_sku: str) -> str:
        """Get product name from KPI data for a given SKU"""
        try:
            product_data = self.kpi_df[self.kpi_df['SKU'] == product_sku]
            if not product_data.empty and 'Name' in product_data.columns:
                name = product_data['Name'].iloc[0]
                return name if pd.notna(name) else ""
            return ""
        except Exception:
            return ""
    
    def _style_worksheet(self, worksheet, dataframe, is_summary=False):
        """
        Apply professional styling to Excel worksheet
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        data_font = Font(size=11)
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        # Set column widths based on sheet type
        if is_summary:
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 25
        else:
            # Auto-size columns based on content
            for col_num, column in enumerate(dataframe.columns, 1):
                col_letter = worksheet.cell(row=1, column=col_num).column_letter
                max_length = max(
                    len(str(column)),  # Header length
                    max([len(str(cell)) for cell in dataframe.iloc[:, col_num-1]], default=0)  # Data length
                )
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 60)  # Cap at 60 chars
        
        # Style header row
        for col_num in range(1, len(dataframe.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Style data rows with alternating colors and spacing
        light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        for row_num in range(2, worksheet.max_row + 1):
            # Alternating row colors
            is_even_row = (row_num % 2 == 0)
            
            for col_num in range(1, len(dataframe.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = border
                
                # Apply alternating background
                if is_even_row:
                    cell.fill = light_fill
                
                # Alignment based on data type
                if col_num == 1:  # SKU columns - left align
                    cell.alignment = left_alignment
                else:  # Numbers - center align
                    cell.alignment = center_alignment
        
        # Increase row height for better readability
        for row_num in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_num].height = 20  # Default is 15

    def generate_reports(self, availability_path: str, replenishment_path: str, bom_path: str, export_excel: bool = True) -> Dict[str, Any]:
        """
        Main method to generate all reports
        Returns dictionary with assembly and transfer recommendations
        """
        try:
            # Load data
            self.load_data(availability_path, replenishment_path, bom_path)
            
            # Generate assembly analysis (returns complete results with summary)
            results = self.analyze_assembly_capacity()
            
            print(f"Report generation completed:")
            print(f"  - {len(results['assembly_ready'])} products ready for assembly")
            print(f"  - {len(results['cannot_assemble'])} products cannot be assembled")
            print(f"  - {len(results['transfer_recommendations'])} transfer recommendations")
            
            # Export to Excel if requested
            if export_excel:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                excel_filename = f"DBI_Assembly_Reports_{timestamp}.xlsx"
                excel_path = self.export_to_excel(results, excel_filename)
                results['excel_file'] = excel_filename
                results['excel_path'] = excel_path
            
            return results
            
        except Exception as e:
            print(f"Error generating reports: {e}")
            raise
